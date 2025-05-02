import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import os
import time

# Load the Excel sheet
if not os.path.exists('D:\Downloades\Programming Folder\test folder\python\general test\attendance.xlsx'):
    print("Error: 'attendance.xlsx' file not found.")
    exit()

book = openpyxl.load_workbook('attendance.xlsx')
sheet = book['Sheet1']

# Check if the sheet has the correct structure
required_columns = ["Roll Number", "Name", "CI Attendance", "Python Attendance", "DM Attendance"]
header = [sheet.cell(row=1, column=i).value for i in range(1, 6)]
if header != required_columns:
    print("Error: 'attendance.xlsx' does not have the required structure.")
    exit()

# Initialize empty cells with 0
r = sheet.max_row
for i in range(2, r + 1):
    for col in range(3, 6):  # Columns for CI, Python, DM attendance
        if sheet.cell(row=i, column=col).value is None:
            sheet.cell(row=i, column=col).value = 0
book.save('attendance.xlsx')

# Email credentials
from_id = os.getenv('EMAIL_ID')
pwd = os.getenv('EMAIL_PASSWORD')
if not from_id or not pwd:
    print("Error: Email credentials not set. Please set EMAIL_ID and EMAIL_PASSWORD as environment variables.")
    exit()

# List of staff mail IDs
staff_mails = ['erakshaya485@gmail.com', 'yyyyyyyy@gmail.com']

# Warning messages
m1 = "Warning!!! You can take only one more day leave for CI class."
m2 = "Warning!!! You can take only one more day leave for Python class."
m3 = "Warning!!! You can take only one more day leave for DM class."


def savefile():
    """Save the Excel file."""
    book.save(r'D:\\attendance.xlsx')
    print("Saved!")


def connect_smtp():
    retries = 3
    for attempt in range(retries):
        try:
            s = smtplib.SMTP('smtp.gmail.com', 587, timeout=120)
            s.starttls()
            s.login(from_id, pwd)
            return s
        except Exception as e:
            print(f"Attempt {attempt + 1} failed: {e}")
            time.sleep(5)
    print("Error: Could not connect to the SMTP server.")
    exit()


def check(no_of_days, row_num, b):
    """Check attendance and send emails."""
    global staff_mails

    l1 = []  # List of students to remind
    l2 = ""  # Concatenated roll numbers with lack of attendance
    l3 = []  # List of roll numbers with lack of attendance

    for student in range(len(row_num)):
        # If total number of leaves equals the threshold
        if no_of_days[student] == 2:
            if b == 1:
                l1.append(sheet.cell(row=row_num[student], column=2).value)
                mailstu(l1, m1)  # Send mail
            elif b == 2:
                l1.append(sheet.cell(row=row_num[student], column=2).value)
                mailstu(l1, m2)
            else:
                l1.append(sheet.cell(row=row_num[student], column=2).value)
                mailstu(l1, m3)

        # If total number of leaves exceeds the threshold
        elif no_of_days[student] > 2:
            l2 += str(sheet.cell(row=row_num[student], column=1).value) + " "
            l3.append(sheet.cell(row=row_num[student], column=2).value)

            subject = "Unknown"  # Default value for subject
            if b == 1:
                subject = "CI"
            elif b == 2:
                subject = "Python"
            elif b == 3:
                subject = "Data Mining"

    # If threshold crossed, modify the message
    if l2 and l3:
        msg1 = f"You have lack of attendance in {subject}!!!"
        msg2 = f"The following students have lack of attendance in your subject: {l2.strip()}"

        mailstu(l3, msg1)  # Mail to students
        staff_id = staff_mails[b - 1]  # Pick respective staff's mail ID
        mailstaff(staff_id, msg2)  # Mail to staff


def mailstu(li, msg):
    """Send email to students."""
    try:
        s = connect_smtp()

        for to_id in li:
            message = MIMEMultipart()
            message['Subject'] = 'Attendance Report'
            message.attach(MIMEText(msg, 'plain'))
            content = message.as_string()
            s.sendmail(from_id, to_id, content)

        s.quit()
        print("Mail sent to students.")
    except Exception as e:
        print(f"Error sending mail to students: {e}")


def mailstaff(mail_id, msg):
    """Send email to staff."""
    try:
        s = connect_smtp()

        message = MIMEMultipart()
        message['Subject'] = 'Lack of Attendance Report'
        message.attach(MIMEText(msg, 'plain'))
        content = message.as_string()
        s.sendmail(from_id, mail_id, content)

        s.quit()
        print("Mail sent to staff.")
    except Exception as e:
        print(f"Error sending mail to staff: {e}")


# Main loop
resp = 1
while resp == 1:
    print("1 ---> CI\n2 ---> Python\n3 ---> DM")

    # Validate subject input
    while True:
        try:
            y = int(input("Enter subject (1 ---> CI, 2 ---> Python, 3 ---> DM): "))
            if y not in [1, 2, 3]:
                raise ValueError
            break
        except ValueError:
            print("Invalid input. Please enter 1, 2, or 3.")

    # Validate number of absentees
    while True:
        try:
            no_of_absentees = int(input("Number of absentees: "))
            if no_of_absentees < 1:
                raise ValueError
            break
        except ValueError:
            print("Invalid input. Please enter a positive integer.")

    # Validate roll numbers
    while True:
        try:
            if no_of_absentees > 1:
                x = list(map(int, input("Roll numbers: ").split()))
            else:
                x = [int(input("Roll number: "))]
            break
        except ValueError:
            print("Invalid input. Please enter valid roll numbers.")

    # Validate roll numbers
    valid_roll_numbers = [sheet.cell(row=i, column=1).value for i in range(2, r + 1)]
    for student in x:
        if student not in valid_roll_numbers:
            print(f"Error: Roll number {student} not found in the attendance sheet.")
            exit()

    # List to hold row of the student in the Excel sheet
    row_num = []

    # List to hold total number of leaves taken by each student
    no_of_days = []

    for student in x:
        for i in range(2, r + 1):
            if y == 1:
                if sheet.cell(row=i, column=1).value == student:
                    m = sheet.cell(row=i, column=3).value
                    m += 1
                    sheet.cell(row=i, column=3).value = m
                    savefile()
                    no_of_days.append(m)
                    row_num.append(i)
            elif y == 2:
                if sheet.cell(row=i, column=1).value == student:
                    m = sheet.cell(row=i, column=4).value
                    m += 1
                    sheet.cell(row=i, column=4).value = m
                    savefile()
                    no_of_days.append(m)
                    row_num.append(i)
            elif y == 3:
                if sheet.cell(row=i, column=1).value == student:
                    m = sheet.cell(row=i, column=5).value
                    m += 1
                    sheet.cell(row=i, column=5).value = m
                    savefile()
                    no_of_days.append(m)
                    row_num.append(i)

    check(no_of_days, row_num, y)
    resp = int(input("Another subject? 1 ---> Yes, 0 ---> No: "))
