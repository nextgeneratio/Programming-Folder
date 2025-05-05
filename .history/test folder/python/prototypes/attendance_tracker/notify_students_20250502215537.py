import openpyxl
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from dotenv import load_dotenv
import os

# Load environment variables from .env file
load_dotenv()
EMAIL_ID = os.getenv('EMAIL_ID')
EMAIL_PASSWORD = os.getenv('EMAIL_PASSWORD')

if not EMAIL_ID or not EMAIL_PASSWORD:
    print("Error: Email credentials are not set in the .env file.")
    exit()

# Load the Excel file
file_path = r'd:\Downloades\Programming Folder\test folder\python\prototypes\attendance_tracker\attendance.xlsx'
try:
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
except FileNotFoundError:
    print(f"Error: File '{file_path}' not found.")
    exit()

# Staff email addresses for each subject
staff_emails = {
    "CI": "staff_ci@example.com",
    "Python": "staff_python@example.com",
    "DM": "staff_dm@example.com"
}

# Function to send email
def send_email(receiver_email, subject, body):
    try:
        # Create the email
        message = MIMEMultipart()
        message['From'] = EMAIL_ID
        message['To'] = receiver_email
        message['Subject'] = subject
        message.attach(MIMEText(body, 'plain'))

        # Connect to the SMTP server and send the email
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(EMAIL_ID, EMAIL_PASSWORD)
            server.sendmail(EMAIL_ID, receiver_email, message.as_string())
        print(f"Email sent to {receiver_email}")
    except Exception as e:
        print(f"Error sending email to {receiver_email}: {e}")

# Iterate through the rows in the Excel sheet
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    print(row)  # Print each row to check its structure

    # Ensure the row has at least 6 columns
    if len(row) < 6:
        print(f"Skipping row due to insufficient data: {row}")
        continue

    # Unpack the first 6 columns
    try:
        roll_number, name, ci_marks, python_marks, dm_marks, student_email = row[:6]
    except ValueError:
        print(f"Skipping row due to unexpected structure: {row}")
        continue

    # Validate email address
    if not isinstance(student_email, str) or "@" not in student_email:
        print(f"Skipping row due to invalid email address: {row}")
        continue

    # Check for marks below 40
    subjects = {"CI": ci_marks, "Python": python_marks, "DM": dm_marks}
    for subject, marks in subjects.items():
        if marks is not None:
            try:
                marks = int(marks)  # Convert marks to an integer
            except ValueError:
                print(f"Skipping invalid marks value for {name} in {subject}: {marks}")
                continue

            if marks < 40:
                # Notify the student
                student_message = (
                    f"Dear {name},\n\n"
                    f"You have scored {marks} marks in {subject}, which is below the passing threshold of 40.\n"
                    f"Please contact your instructor for further guidance.\n\n"
                    f"Best regards,\nYour School Administration"
                )
                send_email(student_email, f"Low Marks Alert: {subject}", student_message)

                # Notify the staff
                staff_email = staff_emails.get(subject)
                if staff_email:
                    staff_message = (
                        f"Dear Instructor,\n\n"
                        f"Student {name} (Roll Number: {roll_number}) has scored {marks} marks in {subject}.\n"
                        f"Please take the necessary actions.\n\n"
                        f"Best regards,\nYour School Administration"
                    )
                    send_email(staff_email, f"Low Marks Alert: {subject}", staff_message)